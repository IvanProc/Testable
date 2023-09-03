import numpy
import sklearn.metrics
from lxml import etree
import ast
import time
import numpy as np
import re
import collections
import logging
import unittest
import calculator
import socket
import time
import timeit
from openpyxl import Workbook
import openpyxl
from PyPDF4 import PdfFileReader, PdfFileWriter
from fpdf import FPDF, XPos, YPos
from collections import deque
import datetime
import requests
from bs4 import BeautifulSoup
import pandas as pd
import matplotlib.pyplot as plt
import scipy
from scipy import integrate
from sklearn.dummy import DummyRegressor
from nltk.corpus import stopwords
# nltk.download('stopwords')
from nltk import edit_distance
from sklearn.feature_extraction.text import CountVectorizer
from nltk import sent_tokenize
import nltk
from nltk.corpus import brown
from nltk.corpus import conll2000
from nltk.corpus import gutenberg
from nltk.corpus import movie_reviews
from nltk.util import ngrams
from scipy import stats
from vega_datasets import data
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
import squarify
import seaborn as sb
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from nltk import ngrams
# nltk.download('punkt')
from nltk.collocations import *
from nltk import word_tokenize
from nltk.probability import FreqDist
from astropy.cosmology import FlatLambdaCDM
from astropy import units as u
from sklearn.metrics import mean_absolute_percentage_error
from sklearn.datasets import load_diabetes
import math
from sklearn import datasets
from sklearn.cluster import KMeans
from sklearn.cluster import DBSCAN
import astropy.units as u
from astropy.constants import mu0
import astropy.constants as const
from astropy.coordinates import SkyCoord
from sklearn.datasets import fetch_california_housing
from sklearn.preprocessing import RobustScaler
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import PowerTransformer
import gensim
from pprint import pprint
import statistics
from sklearn.datasets import load_breast_cancer
from sklearn.model_selection import GridSearchCV
from sklearn.svm import SVC
from scipy.stats import loguniform
from sklearn.model_selection import RandomizedSearchCV
from sklearn.pipeline import make_pipeline

X, y = load_breast_cancer(return_X_y=True)
scaler = StandardScaler()
X = scaler.fit_transform(X)

kernel = ['linear', 'rbf', 'poly']
gamma = [0.1, 1, 10, 100]
C = [0.1, 1, 10, 100, 1000]

param_grid = {
    "kernel": kernel,
    "gamma": gamma,
    "C": C,
}

grid_search = GridSearchCV(
    estimator=SVC(), param_grid=param_grid, n_jobs=-1
)
grid_search.fit(X, y)
best_params = grid_search.best_params_
print("BP")
print(best_params)

param_grid = {
    "kernel": kernel,
    "gamma": loguniform(0.01, 1),
    "C": loguniform(0.01, 1),
}

random_search = RandomizedSearchCV(
    estimator=SVC(),
    param_distributions=param_grid,
    n_iter=10,
    random_state=42,
    n_jobs=-1,
)

random_search.fit(X, y)
best_params = random_search.best_params_

print("BP2")
print(best_params)

# lda = gensim.models.ldamodel.LdaModel.load('C:\\Users\\HVIII\\Downloads\\lda_tripadviser\\lda_model')
# pprint(lda.print_topics())

filename = 'F:\облачные файлы\hyperskill-dataset-85106864.txt'
f = open(filename)
lst = f.read().split()  # now it contains points as strings
f.close()

new_list = []  # create an empty list for adding integers
for i in lst:
    new_list.append(int(i)) # transform a string representation into an integer and append it

print("LALA=")
print(statistics.mean(new_list) + new_list[len(new_list) - 1])

print(bool(2) == 1)

X = [[ 8, -2, 3], [ 2, 25, 0 ], [ 0, 0, -2]]
mms = StandardScaler()
df_mms = mms.fit(X)
# df_mms = pd.DataFrame(df_mms)
print(df_mms.mean_)

data = fetch_california_housing(as_frame=True)
df = data.frame[['HouseAge']]
print(df.head)
print(df.describe())
power_trnsfrmr = PowerTransformer()
df_pt = power_trnsfrmr.fit_transform(df)
df_pt = pd.DataFrame(df_pt, columns=df.columns)
print('answer')
print(df_pt.describe())

lengths = np.array([49, 58, 947])
print(lengths * u.kilometer)
time = 14 * u.hour
print(time.to(u.second))
print(mu0.value)
print(mu0)
polaris = SkyCoord(ra=37.95456067*u.degree, dec=89.26410897*u.degree, frame='icrs')
aldebaran = SkyCoord.from_name('aldebaran')
print(aldebaran)
F = const.G * 488 * u.kg
print(F)
polaris = SkyCoord.from_name('polaris')
print(polaris.separation(aldebaran))

iris = datasets.load_iris()
X = iris.data[:, 2:]
y = iris.target
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.33, random_state=42)

plt.scatter(X_train[:, 0], X_train[:, 1], c=y_train, cmap=plt.cm.tab20b)
plt.xlabel("Petal length")
plt.ylabel("Petal width")
plt.show()

kmeans = KMeans(n_clusters=3, random_state=0)
kmeans.fit(X_train)
test_labels = kmeans.predict(X_test)

plt.scatter(X_test[:, 0], X_test[:, 1], c=test_labels, cmap=plt.cm.tab20b)
plt.xlabel("Petal length")
plt.ylabel("Petal width")
plt.show()

diabetes = load_diabetes()

X = diabetes.data
y = diabetes.target
model = LinearRegression()
X_train, X_test = X[:-50], X[-50:]
y_train, y_test = y[:-50], y[-50:]
print(y_test)
model.fit(X_train, y_train)
y_pred = model.predict(X_test)
print('mse', math.ceil(sklearn.metrics.mean_squared_error(y_test, y_pred)))




array = np.array([[3, 0], [-1, 2]])
u, s, v = np.linalg.svd(array)
print(u)

data = pd.read_csv('data.csv')
dataset = ['rating', 'draft_round', 'bmi']
X, y = data[dataset], data['salary']
X_train, X_test, y_train, y_test = train_test_split(X, y, train_size=0.7, random_state=100)
model = LinearRegression()
model.fit(X_train, y_train)
y_predicted = model.predict(X_test)
mape = mean_absolute_percentage_error(y_test, y_predicted)
y_non_negative = [max(y, 0) for y in y_predicted]
mape_non_negative = mean_absolute_percentage_error(y_test, y_non_negative)
y_median_substitution = []
for y in y_predicted:
    if y < 0:
        y_median_substitution.append(np.median(y_test))
    else:
        y_median_substitution.append(y)
mape_median_substitution = mean_absolute_percentage_error(y_test, y_median_substitution)
print(round(min(mape, mape_non_negative, mape_median_substitution), 5))



plt.scatter(data['rating'], data['salary'])
plt.title('one-variable linear regression')
plt.xlabel('rating')
plt.ylabel('salary')
plt.show()



general = pd.read_csv('general.csv')
prenatal = pd.read_csv('prenatal.csv')
sports = pd.read_csv('sports.csv')

prenatal.columns = [*general.columns]
sports.columns = [*general.columns]

hospitals = pd.concat([general, prenatal, sports], ignore_index=True)
hospitals = hospitals.drop(columns=hospitals.columns[0])
hospitals = hospitals.dropna(how='all')
hospitals['gender'] = hospitals['gender'].replace(['female', 'woman'], 'f').replace(['male', 'man'], 'm')
hospitals['gender'].loc[hospitals['hospital'] == 'prenatal'] = hospitals['gender'].loc[hospitals['hospital'] == 'prenatal'].fillna('f')
hospitals[['bmi', 'diagnosis', 'blood_test', 'ecg', 'ultrasound', 'mri', 'xray', 'children', 'months']] = \
    hospitals[['bmi', 'diagnosis', 'blood_test', 'ecg', 'ultrasound', 'mri', 'xray', 'children', 'months']].fillna(0)

biggest_hospital = hospitals['hospital'].value_counts().index[hospitals['hospital'].value_counts()
                                               == hospitals['hospital'].value_counts().max()][0]
stomach_share = round(hospitals.loc[hospitals.hospital == 'general'].loc[hospitals.diagnosis == 'stomach'].shape[0]
                 / hospitals.loc[hospitals.hospital == 'general'].shape[0], 3)
dislocation_share = round(hospitals.loc[hospitals.hospital == 'sports'].loc[hospitals.diagnosis == 'dislocation'].shape[0]
                 / hospitals.loc[hospitals.hospital == 'sports'].shape[0], 3)
age_difference = int(hospitals.loc[hospitals.hospital == 'general', 'age'].median()
                       - hospitals.loc[hospitals.hospital == 'sports', 'age'].median())
blood_test = hospitals.loc[hospitals.blood_test == 't']
biggest_bt_hospital = blood_test['hospital'].value_counts().index[blood_test['hospital'].value_counts()
                                               == blood_test['hospital'].value_counts().max()][0]
bt_qty = blood_test['hospital'].value_counts().max()
print(f'''The answer to the 1st question is {biggest_hospital}
The answer to the 2nd question is {stomach_share}
The answer to the 3rd question is {dislocation_share}
The answer to the 4th question is {age_difference}
The answer to the 5th question is {biggest_bt_hospital}, {bt_qty} blood tests''')


plt.figure()
plt.hist(hospitals['age'], bins=[0, 15, 35, 55, 70, 80])
plt.figure()
plt.pie(hospitals['diagnosis'].value_counts(), labels=hospitals['diagnosis'].value_counts().index, autopct='%.0f%%')
plt.figure()
plt.violinplot(hospitals['height'])
plt.show()




print(hospitals.shape)
print(hospitals.sample(n=20, random_state=30))





df = pd.read_csv('https://stepik.org/media/attachments/lesson/665342/galaxies_coordinates.tsv', sep='\t')
print(df.head())
properties = [76.98939987756634, 0.6165452599525452, 0.899420976638794, 0.17506762692490038]

my_cosmo = FlatLambdaCDM(H0=67.74, Om0=0.3089)
z = 0.06
angular_diameter_distance = my_cosmo.angular_diameter_distance(z).to(u.kpc)
print('add =', angular_diameter_distance)
for prop in properties:
    print(prop, end=' ')




apartment = '<?xml version="1.0" encoding="utf-8"?><apartment><living_room><furniture><item number="5">chair</item><item number="1">sofa</item><item number="2">table</item><item number="3">bookcase</item><item number="2">armchair</item></furniture><other><item number="1">fireplace</item></other></living_room><bedroom><furniture><item number="1">bed</item><item number="2">bedside_cabinet</item></furniture><other><item number="2">lamp</item></other></bedroom></apartment>'

soup = BeautifulSoup(apartment, 'xml')
print(soup.find('living_room'))






stop_words = set(stopwords.words('english'))
words_doc = nltk.Text(gutenberg.words('blake-poems.txt'))
words_doc = [word.lower() for word in words_doc if word.isalpha()]
words_doc = [word for word in words_doc if word not in stop_words]
freqd = FreqDist()
for word in words_doc:
    freqd[word] += 1

freqd.plot(30)



sentence = 'Roman victory in the Punic Wars and Macedonian Wars established Rome as a super power'

finder = BigramCollocationFinder.from_words(word_tokenize(sentence))
bigrams = nltk.collocations.BigramAssocMeasures()

print(finder.nbest(bigrams.pmi, 5))



fig = make_subplots(rows=1, cols=2)

fig.add_trace(
    go.Bar(x=["Mon", "Tue", "Wed", "Thu", "Fri"], y=[17, 30, 12, 15, 16]),
    row=1, col=1
)

fig.add_trace(
    go.Bar(x=["Sat", "Sun"], y=[30, 25]),
    row=1, col=2
)

fig.update_layout(height=600, width=800, title_text="Sales of this week")
fig.show()


penguins = sb.load_dataset('penguins')
sb.set_style('whitegrid')
sb.set_palette('Accent')

plot = sb.scatterplot(x='flipper_length_mm', y='body_mass_g', data=penguins, hue='sex')
plot.set(xlabel='flipper length', ylabel='body mass', title='Penguins body mass and flipper length')
sb.despine()
sb.pairplot(hue='species', data=penguins)


transportation_models = {
    "WALK": 23, "BIKE": 11, "CAR": 15,
    "TRAM": 12, "BUS": 8, "TRAIN": 12
}

models = list(transportation_models.keys())
number_of_people = list(transportation_models.values())
plt.figure(figsize=(10, 6))
squarify.plot(
    sizes=number_of_people,
    label=models,
    value=number_of_people,
    color=['g', 'r', 'b', 'purple', 'yellow'])
plt.title("Number of people who use each transportation model", fontsize=17)
plt.axis("off")
plt.show()


months = range(1, 13)
bonnie_carrots = [14, 13, 10, 15, 17, 15, 15, 13, 12, 10, 14, 11]
clyde_carrots = [13, 17, 12, 11, 11, 10, 15, 14, 13, 12, 11, 15]
plt.plot(months, bonnie_carrots, color='purple', label='Bonnie')
plt.plot(months, clyde_carrots, color="blue", label="Clyde")
plt.fill_between(months, bonnie_carrots, clyde_carrots, color="darkorange")
# plt.stackplot(months, bonnie_carrots, clyde_carrots, colors=['darkorange', 'purple'], labels=["Bonnie", "Clyde"])
plt.xlabel("Months")
plt.ylabel("Number of carrots")
plt.title("Bonnie and Clyde's monthly carrot intake")
plt.xticks(range(1, 13))
plt.yticks(range(0, 31, 5))
plt.grid()
plt.legend()
plt.show()



years = ['2016', '2017', '2018', '2019']
cats = [50, 45, 37, 30]
dogs = [40, 39, 50, 55]
hamsters = [10, 16, 13, 15]

plt.figure(figsize=(10, 6))

plt.bar(years, cats, label='Cats')
plt.bar(years, dogs, bottom=cats, label='Dogs')
plt.bar(years, hamsters, bottom=[x + y for x, y in zip(cats, dogs)], label='Hamsters')

plt.xlabel('Years', fontsize=14)
plt.ylabel('Preference (%)', fontsize=14)
plt.title('The results of cat/dog survey', fontsize=20)
plt.legend()

plt.show()

years = ["2016", "2017", "2018", "2019"]
cats = [57, 50, 47, 30]
dogs = [43, 50, 53, 70]

x_axis = list(range(len(years)))
plt.figure(figsize=(10, 5))
plt.bar([x - 0.2 for x in x_axis], cats, width=0.3, label='cats', alpha=0.5)
plt.bar(x_axis, dogs, width=0.3, bottom=cats, label='dogs', alpha=0.5)
plt.xticks(x_axis, years)
plt.title('The results of cat/dog survey', fontsize=18)
plt.xlabel('Years', fontsize=12)
plt.ylabel('Preference (%)', fontsize=12)
plt.legend()
plt.show()



films = ['Wonder Woman', 'Sonic', '1917', 'Star Wars', 'Onward']
box_office = [16.7, 26.1, 37.0, 34.5, 10.6]
plt.bar(films, box_office)
plt.xlabel('Film title')
plt.ylabel('Box office')
plt.title('Box office of 5 different films of 2020 in the USA')
plt.grid(color='purple', axis='y', linestyle=':', linewidth=0.9, alpha=0.5)
plt.show()


'''grouped_galaxies = pd.read_csv('https://stepik.org/media/attachments/lesson/665342/galaxies_morphology.tsv', sep='\t')
grouped_galaxies = grouped_galaxies.groupby(['Group'], as_index=False).agg({'n': 'mean', 'T': 'mean'})
df = pd.read_csv('https://stepik.org/media/attachments/lesson/665342/groups.tsv', sep='\t')
df = df.dropna()
grouped_galaxies = grouped_galaxies.merge(df)
plt.scatter(list(grouped_galaxies.T), grouped_galaxies.mean_mu)
plt.scatter(list(grouped_galaxies.n), grouped_galaxies.mean_mu)
plt.xlabel('<T>, <n>')
plt.ylabel('mean_mu')
plt.legend(['mean_T', 'mean_n'])
plt.show()
print(grouped_galaxies.head(20))
shapiro_mu = stats.shapiro(grouped_galaxies.mean_mu).pvalue
shapiro_n = stats.shapiro(grouped_galaxies['n']).pvalue
shapiro_t = stats.shapiro(grouped_galaxies['T']).pvalue
pearson_mu_n = stats.pearsonr(list(grouped_galaxies.mean_mu), list(grouped_galaxies.n)).pvalue
pearson_mu_t = stats.pearsonr(list(grouped_galaxies.mean_mu), list(grouped_galaxies['T'])).pvalue
print(shapiro_mu, shapiro_n, shapiro_t, pearson_mu_n, pearson_mu_t)'''


weather = data.seattle_weather()
weather = weather[['precipitation', 'temp_max', 'temp_min', 'wind']]
print(weather.corr())
print(weather.head())
corr = weather.corr()

plt.imshow(corr, cmap='coolwarm')
plt.colorbar()
plt.title('Weather in Seattle', fontsize=20, pad=20)
plt.gcf().set_size_inches(8, 8)
ticks = ['precipitation', 'temp_max', 'temp_min', 'wind']
plt.xticks(range(len(weather.columns)), ticks, fontsize=10, rotation=15)
plt.yticks(range(len(weather.columns)), ticks, fontsize=10, rotation=15)
labels = corr.values

for y in range(labels.shape[0]):
    for x in range(labels.shape[1]):
        plt.text(y, x, '{:.2f}'.format(labels[x, y]), va='center', ha='center', fontsize=10, rotation=15, color='purple')

data1 = [0, 20, 20, 20, 20, 20, 70]
data2 = [0, 20, 50, 50, 50, 90, 50]
fig, axes = plt.subplots()
axes.set_xticks([1, 2])
axes.set_xticklabels(['Jan', 'Aug'])
axes.set_ylabel('Percentage')
axes.set_title('Violin')
graph = plt.violinplot([data1, data2], showmeans=True, showmedians=True, quantiles=[[0.1, 0.9], [0.1, 0.9]])
graph['cmeans'].set_color('purple')
graph['cmedians'].set_color('k')
graph['cquantiles'].set_color('red')
plt.show()


months=["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug","Sep", "Oct", "Nov", "Dec"]
avg_temperature = [8, 12, 15, 20, 21, 23, 24, 23, 23, 20, 14, 11]
avg_rainfall = [110, 105, 94, 61, 66, 31, 22, 23, 98, 110, 188, 143]
#colors
snow = "lightblue"
rain = "lightgray"
sun = "yellow"
hot = "orange"
skye = [snow, rain, rain, sun, sun, hot, hot, hot, rain, rain, snow, snow]
plt.scatter(months, avg_temperature, s=[2 * n for n in avg_rainfall], c=skye, alpha=0.6)
plt.xlabel("Months")
plt.ylabel("Average temperature")
plt.show()

travel = ['flight', 'car', 'train', 'taxi']
price = [142, 62, 36, 100]
plt.title("Travel Costs in Euro: Vienna - Budapest")
plt.xlabel("Transportation")
plt.ylabel("Price in Euro")
plt.scatter(travel, price, c=price, s=100, cmap='viridis', marker='x')
plt.colorbar()
plt.show()



'''plt.hist([iso_list, group_list], bins=24, label=['iso', 'group'])
plt.legend()

df4 = pd.read_csv('hyperskill-dataset-75601068.txt')
df4.totsp = df4.livingsp + df4.nonlivingsp
print(df4.totsp.sum())
print(df4.head())'''

with open('hyperskill-dataset-75528283.txt') as file:
    sentences = sent_tokenize(file.read())
    vectorizer = CountVectorizer(ngram_range=(2, 3))
    X = vectorizer.fit_transform(sentences)
    print(list(vectorizer.get_feature_names_out()))

text = 'You can use the following snippet to get n-grams for a sentence'
unigram_model = list(ngrams(text.split(), n=1))
bigram_model = list(ngrams(text.split(), n=2))
print(unigram_model)
print(bigram_model)


 # row number of a sentence in Movie reviews corpus
print(movie_reviews.sents(categories='neg')[0])

print(gutenberg.fileids())

print(conll2000.chunked_sents()[1])

print(brown.tagged_words()[:5])


keyboard_instruments = pd.DataFrame({'cat_id': ['001', '002', '003'],
                                     'Instrument': ['Acoustic piano', 'Electric piano', 'Synthesizer'],
                                     'Average price': ['$10,000', '$5,000', '$1,200']},
                                    index=[1, 2, 3])

string_instruments = pd.DataFrame({'cat_id': ['004', '005', '006'],
                                   'Instrument': ['Acoustic guitar', 'Cello', 'Violin'],
                                   'Average price': ['$2,000', '$1,500', '$2,000']},
                                  index=[1, 2, 3])

print(pd.concat([keyboard_instruments, string_instruments], ignore_index=True))

string1 = 'tomato'
string2 = 'potato'
print('distance:', edit_distance('Levenshtein', 'Lenvevstein'))
print('distance:', edit_distance('Levenshtein', 'Levewsthein'))
print('distance:', edit_distance('Levenshtein', 'Lewenhstein'))
print('distance:', edit_distance('Levenshtein', 'Lenevsthein'))
print('distance:', edit_distance('Levenshtein', 'Lewehstein'))

string3 = 'fruit'
string4 = 'frustrated'
print('distance:', nltk.edit_distance(string3, string4))  # 6

data = [163, 163, 164, 170, 170, 172, 173, 190]
andy_data = [161, 172, 174, 175, 181, 183, 186, 190]
plt.hist([data, andy_data], color=['purple', 'black'], edgecolor='orange', stacked=False, label=['mine', 'Andy\'s'],
         bins=list(range(160, 196, 4)))
plt.xlabel('friends')
plt.ylabel('height')
plt.legend()
plt.title('friend\'s heights')

reviews = ['easily the best album of the year.', 'the album is amazing.', "loved the clean production!"]
stopwords = stopwords.words('english')
vectorizer = CountVectorizer(ngram_range=(2,2), binary=False, stop_words=stopwords)
X = vectorizer.fit_transform(reviews)
print(vectorizer.get_feature_names_out())
print(X.toarray())


'''dummy_regressor = DummyRegressor(strategy='quantile', quantile=0.4)
df = pd.read_csv('hyperskill-dataset-75391137.txt')
# df = pd.read_csv('test.txt')
X, y = df["X"], df["y"]
dummy_regressor.fit(X, y)
print(y)
print(dummy_regressor.predict(X)[0])'''



df = pd.read_csv('https://stepik.org/media/attachments/lesson/665342/groups.tsv', sep='\t')
df = df.dropna()
print(df.head())
df_lsb = df[df.features == 1]
df_no_lsb = df[df.features == 0]
print('PVALUE')
print(stats.shapiro(df_lsb.mean_mu).pvalue)
print(stats.shapiro(df_no_lsb.mean_mu).pvalue)
print(stats.fligner(df_lsb.mean_mu, df_no_lsb.mean_mu).pvalue, stats.f_oneway(df_lsb.mean_mu, df_no_lsb.mean_mu).pvalue)
fig, (ax1, ax2) = plt.subplots(1, 2)
ax1.boxplot(df.loc[df.features == 1].mean_mu, showmeans=True)
ax1.set_title('1')
ax2.set_title('0')
ax2.boxplot(df.loc[df.features == 0].mean_mu, showmeans=True)
print(df.loc[df.features == 1].mean_mu.mean(), df.loc[df.features == 0].mean_mu.mean())
'''df.plot(y=['Family', 'Health (Life Expectancy)', 'Economy (GDP per Capita)'], kind='hist', bins=8, alpha=0.1)
# df['Region'].value_counts().plot(kind='bar')
df.plot(x='Economy (GDP per Capita)', y='Happiness Score', kind='scatter')
df.plot(y=['Family', 'Freedom', 'Trust (Government Corruption)'], kind='box', showmeans=True)

df = pd.read_csv("hyperskill-dataset-75240811.txt", sep=',')
df = df.dropna(axis=1, thresh=8)
for column in df.columns:
    median = df[column].median()
    print(median)
    df[column] = df[column].fillna(median)
print(df.head(5))'''


'''np.random.seed(14)
data_1 = np.random.normal(50, 40, 200)
data_2 = np.random.normal(60, 30, 200)
data_3 = np.random.normal(70, 20, 200)
data_4 = np.random.normal(80, 10, 200)

data = [55, 27, 15, 3]
labels = ['Chocolate', 'Vanilla', 'Strawberry', 'Other']
fig, (ax2, ax3) = plt.subplots(1, 2, figsize=(12, 8))
plt.xlabel("x axis", fontsize=12)
plt.ylabel('y axis', fontsize=12)
explode = [0.08, 0.08, 0.08, 0.08]
colors = ['saddlebrown', 'wheat', 'crimson', 'purple']
ax2.pie(data, radius= 1., labels=labels, explode=explode, colors=colors, counterclock=True, shadow=True, autopct='%.0f%%',
        startangle=45, wedgeprops={'width': 0.4, 'alpha': 0.7})
ax3.pie(data, labels=labels, explode=explode, colors=colors, counterclock=False, shadow=True, autopct='%.0f%%',
        startangle=45, wedgeprops={'alpha': 0.7})
ax2.set_title('counterclockwise', fontsize=8)
ax3.set_title('clockwise', fontsize=8)
ax2.set_ylabel('ax1')
ax3.set_ylabel('ax2')
# plt.title('The results of the icecream survey', fontsize=10)
plt.legend(labels)

pets_data = [45, 41, 10, 4]
pets_labels = ['Dogs', 'Cats', 'Parrots', 'Other']
pets_colors = ['orange', 'teal', 'powderblue', 'grey']
ax2.pie(pets_data, radius=0.5, labels=pets_labels, labeldistance=0.5, colors=pets_colors,
        wedgeprops={'width': 0.3})


data = [data_1, data_2, data_3, data_4]
plt.figure(figsize=(8, 12))
labels = ['first', 'second', 'third', 'forth']
boxprops = {'facecolor': 'lightblue', 'edgecolor': 'teal', 'linewidth': 2.0}
whiskerprops = {'color': 'green', 'linewidth': 1.5}
capprops = {'color': 'orange', 'linewidth': 1.5}
medianprops = {'color': 'black', 'linewidth': 2}
plt.ylabel('Values')
plt.xlabel('Dataset')
plt.title('Multiple plot example')
plt.boxplot(data, vert=True, labels=labels, patch_artist=True, boxprops=boxprops, whiskerprops=whiskerprops,
            capprops=capprops, meanprops=medianprops)'''


'''x = np.linspace(0, 10)
y = x

fig1 = plt.figure(1)

plt.plot(x,y)'''


fig, axes = plt.subplots(1, 2, figsize=(10, 5))
ax1, ax2 = axes
fig.suptitle("My pretty plot", fontsize=30)

x = np.linspace(0, 10, 200)
y1 = np.sin(x) + x
y2 = np.cos(x)

color = 'r'
linestyle = 'dotted'
linewidth = 4

ax1.set_xlabel('first arg', fontsize=16)
ax1.set_ylabel('sin(x) + x', fontsize=16)
ax1.plot(x, y1, c=color, linestyle=linestyle, linewidth=linewidth)

ax2.set_xlabel('second arg', fontsize=16)
ax2.set_ylabel('cos(x)', fontsize=16)
ax2.plot(x, y2, c='purple', linewidth=3)

fig.tight_layout(pad=2)

for xtick in ax1.xaxis.get_major_ticks():
    xtick.label1.set_fontsize(8)
for ytick in ax1.yaxis.get_major_ticks():
    ytick.label1.set_fontsize(12)

plt.show()
fig.clf()
plt.close()



"""office_a = pd.DataFrame(pd.read_xml("https://www.dropbox.com/s/jpeknyzx57c4jb2/A_office_data.xml?dl=1"))
office_a.index = ["A" + str(eoid) for eoid in office_a["employee_office_id"]]

office_b = pd.DataFrame(pd.read_xml("https://www.dropbox.com/s/hea0tbhir64u9t5/B_office_data.xml?dl=1"))
office_b.index = ["B" + str(eoid) for eoid in office_b["employee_office_id"]]

hr = pd.DataFrame(pd.read_xml("https://www.dropbox.com/s/u6jzqqg1byajy0s/hr_data.xml?dl=1"))
hr = hr.set_index("employee_id", drop=False)
print(hr.info)

unified_dataset = pd.concat([office_a, office_b])
it_projects = unified_dataset.query("salary == 'low' & Department == 'IT'").loc[:, "number_project"]
unified_dataset = unified_dataset.merge(hr, how="inner", left_on=unified_dataset.index, right_on=hr.index)
unified_dataset = unified_dataset.set_index("key_0")
unified_dataset.index.name = None
unified_dataset = unified_dataset.drop(columns=['employee_office_id', 'employee_id'])
unified_dataset = unified_dataset.sort_index()

first_table = unified_dataset.pivot_table(index="Department", columns=['left', 'salary'],
                                          values='average_monthly_hours', aggfunc='median')
print(first_table.head())
first_table = first_table.loc[(first_table[(0, 'high')] < first_table[0, 'medium']) |
                              (first_table[(1, 'low')] < first_table[1, 'high'])]
print(first_table.to_dict())

second_table = unified_dataset.pivot_table(index='time_spend_company', columns='promotion_last_5years',
                                           values=['satisfaction_level', 'last_evaluation'],
                                           aggfunc=['min', 'max', 'mean']).round(2)
# print(second_table.head(10))
second_table = second_table.loc[second_table[('mean', 'last_evaluation', 0)] > second_table[('mean', 'last_evaluation', 1)]]
print(second_table.head(10))
# top_ten = unified_dataset.sort_values("average_monthly_hours", ascending=False).iloc[0:10, 5]
# projects_number = sum(unified_dataset.query("salary == 'low' & Department == 'IT'").loc[:, "number_project"])
# three_employees_info = unified_dataset.loc[['A4', 'B7064', 'A3033'], ['last_evaluation', 'satisfaction_level']]

def count_five_plus(series):
    count = 0
    for num in series:
        if num > 5:
            count += 1
    return count

boss_table = unified_dataset.groupby('left')\
    .agg({'number_project': ['median', count_five_plus], 'time_spend_company': ['mean', 'median'],
          'Work_accident': 'mean', 'last_evaluation': ['mean', 'std']})

print(boss_table.to_dict())

# print(top_ten.tolist())
# print(projects_number)
# print(three_employees_info.values.tolist())

print(unified_dataset.index.tolist())
print(unified_dataset.columns.tolist())
print(unified_dataset.head())
print(unified_dataset.tail())
print(unified_dataset.info())

ages_list = [21, 20, 25, 22]
names_list = ['Anna', 'Bob', 'Maria', 'Jack']
ages_series = pd.Series(ages_list, index=names_list, name="Age")
print(ages_series)


file = open("my_file_1.xml", "r").read()
soup = BeautifulSoup(file, "xml")
print(soup.prettify())
tag_1 = soup.find("title")
tag_2 = soup.find_all("director")
print(tag_1, tag_2)
tag_3 = soup.find("title", {"year": "2001"})
print(tag_3)
print(soup.director)
print(tag_3.parent)
print(list(soup.find("movie").children))
print(soup.find("movie").contents)
print(list(soup.find("movie").previous_siblings))
print(list(soup.find("movie").next_siblings))
for tag in tag_2:
    print(tag.text)
print(tag_3.get("year"))


print(datetime.datetime.now().time())
print(datetime.datetime.today().tzname())

date_string = "06/04/2020 12.30"
date = datetime.datetime.strptime(date_string, "%d/%m/%Y %H.%M")
print(date)
new_date = date.strftime("%B %d %Y at %H.%M")
print(new_date)

dt4 = datetime.datetime(2020, 6, 4, 0, 30)
date_string3 = dt4.strftime("test %Y-%m-%d-%I:%M")
print(date_string3)

date_string = "06/04/2020 12:30"
dt3 = datetime.datetime.strptime(date_string, "%d/%m/%Y %I:%M")
print(dt3)

stack = deque()
action = "BUY Anna Karenina".split(" ", 1)

pdf_f = FPDF(orientation="L", unit="in")
pdf_f.add_page()
pdf_f.set_font(family="times", style="I", size=14)
pdf_f.cell(txt="Hi, PDF!")
pdf_f.cell(h=1, txt="I am the first line at the left", border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
pdf_f.cell(h=1, txt="I am the centered line", border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
pdf_f.output("hi_pdf.pdf")"""


"""with open("dummy.pdf", "rb") as file:
    pdf = PdfFileReader(file)
    info = pdf.getDocumentInfo()
    page_nums = pdf.numPages"""

metadata = """
author: {info.author}
creator: {info.creator}
producer: {info.producer}
subject: {info.subject}
title: {info.title}
numpages: {page_nums}"""

# print(metadata)

pdf_page = """pdf = PdfFileReader("dummy.pdf")
new_pdf = PdfFileWriter()
page_0 = pdf.getPage(0).rotateClockwise(90)
newest_pdf = PdfFileReader("dummy.pdf")
page_1 = newest_pdf.getPage(0).rotateClockwise(270)
new_pdf.addPage(page_0)
new_pdf.addPage(page_1)
with open("rotated_pdf.pdf", "wb") as new_file:
    new_pdf.write(new_file)"""



a = """book = Workbook()
sheet = book.active
sheet['A1'] = "Katia"
sheet["A2"] = "Barsik"
sheet["A3"] = 'Marsik'
sheet['a4'] = 'Larsik I'
sheet.cell(row=5, column=1).value = 'Pliushik'
book.save("book.xlsx")"""

a3 = """workbook = openpyxl.load_workbook("book.xlsx")
all_sheets = workbook.sheetnames
sheet = workbook[all_sheets[0]]
print(sheet)
sheet = workbook["Sheet"]
print(sheet['a1'].value)
workbook.create_sheet("The last cat")
the_last_cat_sheet = workbook["The last cat"]
the_last_cat_sheet['a1'] = "Barsik II"
workbook.save("book.xlsx")"""

"""def get_odds(n):
    for i in range(n):
        if i % 2 != 0:
            yield i


# Creating generator objects
start_time = time.time()
odd_generator = get_odds(10000000)
end_time = time.time()
print(end_time - start_time)  # 0.000015

# You can iterate over a generator by using a for loop
for _ in range(10):
    print(next(odd_generator))


snippet = "'pineapples and bananas'.split()"
print(timeit.timeit(stmt=snippet, number=1000000))  # returns: 0.11186270000007426



array = np.array([1, 2])
print(type(array) == numpy.ndarray)
print(time.localtime(0))
print(time.asctime())
print(time.ctime(time.time()))
print(time.timezone / 3600)
print(time.daylight)

eugene = set("Greece Netherlands Colombia UK".split())
rose = set("Italy UK Russia Greece Canada".split())
print(eugene.symmetric_difference(rose))

# client_socket = socket.socket()
hostname = "google.com"
port = 8080
address = (hostname, port)
# client_socket.connect(address)

data = "wake up, neo"
data = data.encode()
# client_socket.sendall(data)

# response = client_socket.recv(1024)
# response = response.decode()
# print(response)
# socket.close


logging.basicConfig(format='%(levelname)s:%(message)s', level='DEBUG')
logging.info("Your program is running excellent")

logging.warning("Your %s was executed successfully, but the %s is wrong!", "script", "output")

logging.critical("It is critical to understand logs!")
logging.error("Running this line will result in an error message!")
logging.warning("You must catch that bug! It is a warning!")
logging.info("My info is that you are here!")
logging.debug("I'm debugging!")

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
console_handler = logging.StreamHandler()
log_format = "%(asctime)s | %(levelname)s: %(message)s"
console_handler.setFormatter(logging.Formatter(log_format))
logger.addHandler(console_handler)

logger.debug('Here you have some information for debugging.')
logger.info('Everything is OK. Keep going!')
logger.warning("Something strange has happened, but it's not critical.")
logger.error('Something unexpected and critical has happened.')
logger.critical('A critical error! The code cannot run!')


phrase = [('A', 'Away'), ('F', 'From'), ('K', 'Keyboard')]
unzipped = zip(*phrase)
for letter in unzipped:
    print(letter)
print(list(unzipped))

print("check", isinstance(([0],), collections.abc.Hashable))
# print("check-2", hash(([0],)))

print(("a", "b") in ("b", "a"))

def generator2():
    for x in range(10):
        yield x

def generator3():
    for y in range(100, 110):
        yield y

def generator():
    yield from generator2()
    for y in generator3():
        yield y
    print("stoppp")

for x in generator():
    print(x)


numbers = [1, 2, 3]

my_generator = (n ** 2 for n in numbers)

print(next(my_generator))
# Outputs 1

print(next(my_generator))
# Outputs 4

pets = "Dogs"
# your code here
pattern = re.compile("dog|cat|parrot|hamster", flags=re.I)
print(pattern.findall(pets))


first_np_array = np.array([1, 2, 3, 4, 5])
print(first_np_array)
print(type(first_np_array))

second = np.array([[1, 1, 1],
                   [2, 2, 2]])
print(second)

list_a = [1, 2, 3, 4]
array_a = np.array(list_a)

list_b = [11, 22, 33, 44]
array_b = np.array(list_b)
print(list_a + array_a)
print(array_a + list_a)

first = np.array([1, 2, 3, 4, 5])
second = np.array([[1, 1, 1],
                   [2, 2, 2]])
print(first.shape)
print(first.size)
print(second.shape)
print(second.size)
print(second.ndim)
print(len(second))

array = np.array([[[1, 1, 1], [2, 2, 2]],
                   [[3, 3, 3], [4, 4, 4]]])
print(f"Shape: {array.shape}; dimensions: {array.ndim}; length: {len(array)}; size: {array.size}")

action = "to feed the cat"
c_time = time.strftime("%H:%M:%S", time.localtime())
print(f"It's {c_time}. Time {action}.")


xml_string = '<profile><account login="login" password="secret"/></profile>'
root = etree.fromstring(xml_string)

expression = "expr = 1 + 2
expr_2 = 3 + 4
result = expr * expr_2
print(result, 'This is '
      'result.')"""
"""node = ast.parse(expression)
print(node)
print(ast.dump(node))
for n in node.body:
    print(n, n.lineno, n.end_lineno)
print()
node_new = ast.parse("1 + 2")
nodes = ast.walk(node_new)
for n in nodes:
    print(n)
print()


class BinOpLister(ast.NodeVisitor):
    def visit_BinOp(self, node):
        print(node.left)
        print(node.op)
        print(node.right)
        self.generic_visit(node)


BinOpLister().visit(node_new)

user_input = "15"
print(type(user_input))

check_user_input = ast.literal_eval(user_input)
print(type(check_user_input))


class TestCalculator(unittest.TestCase):

    def setUp(self) -> None:
        self.calc = calculator.Calculator(4, 8)
        print("setUp set on")

    def tearDown(self):
        print('tearDown on')

    def test_add(self):
        self.assertEqual(self.calc.add(), 12)
        print("test add on")

    def test_divide(self):
        with self.assertRaises(ValueError):
            calculator.divide(10, 0)
        # self.assertEqual(calculator.divide(0,0), 0)"""

# if __name__ == "__main__":
#    unittest.main()